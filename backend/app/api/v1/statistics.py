from datetime import datetime, timezone
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from sqlalchemy.orm import selectinload
from typing import List, Dict, Any

from app.db.database import get_db
from app.api.dependencies import RequireRole
from app.models.grading import ExamResult
from app.models.exam import ExamSubmission, ExamParticipant, Exam, ExamForm, ExamFormQuestion
from app.models.question import Question
from app.models.user import User, Role

router = APIRouter()


@router.get("/overview", dependencies=[Depends(RequireRole(["ADMIN", "TEACHER"]))])
async def get_dashboard_overview(
    db: AsyncSession = Depends(get_db),
):
    question_count = await db.scalar(select(func.count(Question.id))) or 0
    exam_count = await db.scalar(select(func.count(Exam.id))) or 0

    participant_count = await db.scalar(
        select(func.count(ExamParticipant.id))
        .join(User, User.id == ExamParticipant.user_id)
        .join(Role, Role.id == User.role_id)
        .where(Role.name == "STUDENT")
    ) or 0

    submission_count = await db.scalar(select(func.count(ExamSubmission.id))) or 0

    recent_result = await db.execute(
        select(Exam).order_by(Exam.created_at.desc()).limit(3)
    )
    recent_exams = [
        {
            "id": exam.id,
            "name": exam.name,
            "start_time": exam.start_time,
            "end_time": exam.end_time,
            "status": exam.status,
        }
        for exam in recent_result.scalars().all()
    ]

    score_result = await db.execute(
        select(ExamResult.raw_total_score, ExamResult.total_score)
        .order_by(ExamResult.created_at.desc())
        .limit(1000)
    )
    scores = [float(raw_score) for raw_score, _ in score_result.all()]
    distribution: List[Dict[str, Any]] = []
    bucket_size = 12
    for bucket in range(10):
        lower = bucket * bucket_size
        upper = (bucket + 1) * bucket_size
        distribution.append({
            "range": f"{lower}-{upper}",
            "count": sum(1 for score in scores if lower <= score < upper),
        })
    if any(score >= 120 for score in scores):
        distribution[-1]["count"] += sum(1 for score in scores if score >= 120)

    return {
        "total_questions": question_count,
        "total_exams": exam_count,
        "total_participants": participant_count,
        "total_submissions": submission_count,
        "recent_exams": recent_exams,
        "score_distribution": distribution,
    }


@router.get("/exams/{exam_id}/export.xlsx", dependencies=[Depends(RequireRole(["ADMIN", "TEACHER"]))])
async def export_exam_results(exam_id: int, db: AsyncSession = Depends(get_db)):
    exam_result = await db.execute(select(Exam).where(Exam.id == exam_id))
    exam = exam_result.scalars().first()
    if not exam:
        raise HTTPException(status_code=404, detail="Exam not found")

    submissions_result = await db.execute(
        select(ExamSubmission)
        .join(ExamParticipant, ExamParticipant.id == ExamSubmission.exam_participant_id)
        .options(
            selectinload(ExamSubmission.participant).selectinload(ExamParticipant.user),
            selectinload(ExamSubmission.participant).selectinload(ExamParticipant.exam_form),
        )
        .where(ExamParticipant.exam_id == exam_id)
        .order_by(ExamSubmission.id)
    )
    submissions = submissions_result.scalars().all()
    result_ids = [submission.id for submission in submissions]
    results_by_submission: dict[int, ExamResult] = {}
    if result_ids:
        results_result = await db.execute(
            select(ExamResult).where(ExamResult.exam_submission_id.in_(result_ids))
        )
        results_by_submission = {
            result.exam_submission_id: result for result in results_result.scalars().all()
        }

    headers = [
        "STT", "SBD", "Họ tên", "Email", "Mã đề", "Điểm mục tiêu",
        "Điểm CTT (0-120)", "Điểm IRT (0-1200)", "Đã chấm", "Điểm phần 1",
        "Điểm phần 2", "Điểm phần 3", "Điểm phần 4",
    ] + [f"C{i}" for i in range(1, 121)]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Kết quả"
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1B45B3")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    for index, submission in enumerate(submissions, start=1):
        participant = submission.participant
        user = participant.user if participant else None
        result = results_by_submission.get(submission.id)
        item_scores = result.item_scores if result and result.item_scores else {}
        ctt_total = result.raw_total_score if result else None
        irt_total = result.total_score if result else None
        row = [
            index,
            participant.sbd if participant else None,
            user.full_name if user and user.full_name else user.username if user else None,
            user.email if user else None,
            participant.exam_form.code if participant and participant.exam_form else None,
            participant.target_score if participant else None,
            ctt_total,
            irt_total,
            bool(result),
            result.ctt_score_part1 if result else None,
            result.ctt_score_part2 if result else None,
            result.ctt_score_part3 if result else None,
            result.ctt_score_part4 if result else None,
        ] + [item_scores.get(str(question_number), -1) for question_number in range(1, 121)]
        sheet.append(row)

    widths = {"A": 8, "B": 14, "C": 26, "D": 30, "E": 12}
    for column_name, width in widths.items():
        sheet.column_dimensions[column_name].width = width
    for column_index in range(6, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(column_index)].width = 13

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"mit-exams-{exam_id}-{datetime.now(timezone.utc):%Y%m%d-%H%M%S}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

@router.get("/exams/{exam_id}/overview", dependencies=[Depends(RequireRole(["ADMIN", "TEACHER"]))])
async def get_exam_overview(exam_id: int, db: AsyncSession = Depends(get_db)):
    # Lấy tổng số lượng nộp bài (join qua ExamParticipant vì ExamSubmission không có exam_id)
    result = await db.execute(
        select(ExamSubmission)
        .join(ExamParticipant, ExamParticipant.id == ExamSubmission.exam_participant_id)
        .where(ExamParticipant.exam_id == exam_id)
    )
    submissions = result.scalars().all()
    
    if not submissions:
        return {
            "total_participants": 0,
            "average_score": 0,
            "max_score": 0,
            "min_score": 0,
            "distribution": []
        }
        
    submission_ids = [s.id for s in submissions]
    
    # Tính điểm
    # Lưu ý: CTT có 4 phần, nếu thi trắc nghiệm ĐGNL tối đa 120 câu, ta có thể dùng sum ctt_score
    r_res = await db.execute(select(ExamResult).where(ExamResult.exam_submission_id.in_(submission_ids)))
    exam_results = r_res.scalars().all()
    
    total_scores = []
    for r in exam_results:
        # Nếu chưa có IRT score, tính tạm raw ctt score
        if r.total_score is not None:
            total_scores.append(r.total_score)
        else:
            total_scores.append(r.ctt_score_part1 + r.ctt_score_part2 + r.ctt_score_part3 + r.ctt_score_part4)
            
    if not total_scores:
        return {"total_participants": len(submissions), "message": "No grading results found"}
        
    avg_score = sum(total_scores) / len(total_scores)
    
    # Tính phổ điểm (distribution)
    # Ví dụ chia 10 bucket: 0-12, 12-24, ... (nếu 120 điểm), hoặc 0-120, 120-240 (nếu 1200 điểm)
    # Ở đây giả định max 120 điểm cho đơn giản
    buckets = {f"{i*12}-{(i+1)*12}": 0 for i in range(10)}
    for score in total_scores:
        idx = int(score // 12)
        if idx >= 10:
            idx = 9
        buckets[f"{idx*12}-{(idx+1)*12}"] += 1
        
    distribution = [{"range": k, "count": v} for k, v in buckets.items()]
    
    return {
        "total_participants": len(total_scores),
        "average_score": round(avg_score, 2),
        "max_score": round(max(total_scores), 2),
        "min_score": round(min(total_scores), 2),
        "distribution": distribution
    }

@router.get("/exams/{exam_id}/items", dependencies=[Depends(RequireRole(["ADMIN", "TEACHER"]))])
async def get_exam_item_analysis(exam_id: int, db: AsyncSession = Depends(get_db)):
    # Tìm các mã đề của kỳ thi
    f_res = await db.execute(select(ExamForm).where(ExamForm.exam_id == exam_id))
    forms = f_res.scalars().all()
    form_ids = [f.id for f in forms]
    
    if not form_ids:
        return []
        
    # Tìm các câu hỏi trong đề
    q_res = await db.execute(
        select(Question).join(ExamFormQuestion, ExamFormQuestion.question_id == Question.id)
        .where(ExamFormQuestion.exam_form_id.in_(form_ids))
        .distinct()
    )
    questions = q_res.scalars().all()
    
    analysis = []
    for q in questions:
        flags = []
        if q.a_param < 0.3:
            flags.append("POOR_DISCRIMINATION")
        if q.b_param > 3.0:
            flags.append("TOO_HARD")
        elif q.b_param < -3.0:
            flags.append("TOO_EASY")
            
        analysis.append({
            "question_id": q.id,
            "content": q.content[:50] + "...",
            "difficulty_b": round(q.b_param, 2),
            "discrimination_a": round(q.a_param, 2),
            "guessing_c": round(q.c_param, 2),
            "is_calibrated": q.is_calibrated,
            "warning_flags": flags
        })
        
    return analysis
